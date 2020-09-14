from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, Side
from PyQt5 import QtGui
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QFileDialog, QHeaderView, QMessageBox,
                             QTableWidget, QTableWidgetItem)

from libs.g import g
from model.area import Area
from ui.page_elements.modal_dialog import ModalDialog
from ui.page_elements.detail_page import DetailPage
from model.zyrs import ZYRS

from .DetailpageUI import Ui_Dialog


class DetailPage(ModalDialog):
    def __init__(self, parent):
        super().__init__(parent, size=(1000, 800))
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        # tableWidget
        table_widget = self.ui.tableWidget
        table_widget.setSelectionMode(QTableWidget.NoSelection)
        # tableWidget-header
        hor_header = self.ui.tableWidget.horizontalHeader()
        hor_header.setSectionResizeMode(QHeaderView.Stretch)
        self.ui.tableWidget.horizontalHeader().setStyleSheet("QHeaderView::section{font:13pt '黑体' ;color: black;};")

        # tableWidget-span
        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsEnabled)
        item.setText("第一选区")
        item.setTextAlignment(Qt.AlignCenter)
        table_widget.setSpan(0, 0, 7, 1)
        table_widget.setItem(0, 0, item)
        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsEnabled)
        item.setText("第二选区")
        item.setTextAlignment(Qt.AlignCenter)
        table_widget.setSpan(7, 0, 7, 1)
        table_widget.setItem(7, 0, item)
        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsEnabled)
        item.setText("第三选区")
        item.setTextAlignment(Qt.AlignCenter)
        table_widget.setSpan(14, 0, 7, 1)
        table_widget.setItem(14, 0, item)
        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsEnabled)
        item.setText("第四选区")
        item.setTextAlignment(Qt.AlignCenter)
        table_widget.setSpan(21, 0, 6, 1)
        table_widget.setItem(21, 0, item)
        # btn-bind
        if g.current_user.permission == 0:
            self.ui.btn_save.hide()
        else:
            self.ui.btn_save.show()
        self.ui.button_ok.clicked.connect(self.close)
        self.ui.btn_save.clicked.connect(self.save)
        self.ui.btn_export.clicked.connect(self.export)
        self.ui.btn_search.clicked.connect(self.search)
        # _init
        header = ['选区', '子选区']
        for i in Area.field:
            header.append(getattr(Area, i).comparator.comment)
        self.ui.tableWidget.setHorizontalHeaderLabels(header)
        # messagebox
        self.message = QMessageBox()
        self.message.setStandardButtons(QMessageBox.Yes)
        self.message.button(QMessageBox.Yes).setText('确认')
        self.reload()

    def search(self):
        dialog = DetailPage(self, ZYRS)
        dialog.set_default_conditions(**self.default_conditions)
        dialog.show_(False, id)

    def export(self):
        filename = QFileDialog.getSaveFileName(self, "选择保存地址", "选区", "excel文件(*.xlsx)")[0]
        if filename == "":
            return
        wb = load_workbook(r"E:\python\election-s-prediction\import_test\xq_test.xlsx")
        ws = wb.active
        fontObj1 = Font(name=u'等线', size=11)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        data = Area.search(page_size=-1)["data"]
        beg = 3
        for index, i in enumerate(data):
            if i['name'] == '地区概况': continue
            ws.cell(beg, 3, value=i.mayor)
            ws.cell(beg, 4, value=i.representative)
            ws.cell(beg, 5, value=i.area_mayor)
            ws.cell(beg, 6, value=i.community)
            ws.cell(beg, 7, value=i.peasant_association)
            ws.cell(beg, 8, value=i.civil_organization)
            ws.cell(beg, 9, value=i.other)
            ws.cell(beg, 3).font = fontObj1
            ws.cell(beg, 4).font = fontObj1
            ws.cell(beg, 5).font = fontObj1
            ws.cell(beg, 6).font = fontObj1
            ws.cell(beg, 7).font = fontObj1
            ws.cell(beg, 8).font = fontObj1
            ws.cell(beg, 9).font = fontObj1
            beg += 1
        try:
            wb.save(filename)
        except PermissionError as e:
            QMessageBox.warning(None, "导出数据", "拒绝访问,请先关闭目标文件")
            return

        QMessageBox.information(None, "导出数据", "导出完毕")

    def save(self):
        tag = 2020
        row = self.ui.tableWidget.rowCount()
        for i in range(row):
            mayor = self.ui.tableWidget.item(i, 2).text()
            area_mayor = self.ui.tableWidget.item(i, 3).text()
            representative = self.ui.tableWidget.item(i, 4).text()
            community = self.ui.tableWidget.item(i, 5).text()
            peasant_association = self.ui.tableWidget.item(i, 6).text()
            civil_organization = self.ui.tableWidget.item(i, 7).text()
            target_area = Area.search(id=i + 1)["data"][0]
            target_area.modify(mayor=mayor, area_mayor=area_mayor, representative=representative, community=community,
                               peasant_association=peasant_association, civil_organization=civil_organization)
        QMessageBox.information(None, "选区", "保存成功!")
        self.reload()

    def additem(self, row, col, text):
        item = QTableWidgetItem()
        if g.current_user.permission == 0:
            item.setFlags(Qt.ItemIsEnabled)
        item.setFont(self.font())
        item.setText(text)
        item.setTextAlignment(Qt.AlignCenter)
        self.ui.tableWidget.setItem(row, col, item)

    def reload(self):
        data = Area.search(page_size=-1)["data"]
        for index, i in enumerate(data):
            self.additem(index, 2, i.mayor)
            self.additem(index, 3, i.representative)
            self.additem(index, 4, i.area_mayor)
            self.additem(index, 5, i.community)
            self.additem(index, 6, i.peasant_association)
            self.additem(index, 7, i.civil_organization)
            self.additem(index, 8, i.other)
